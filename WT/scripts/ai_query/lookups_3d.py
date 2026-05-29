from __future__ import annotations

from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .coverage import validate_query_inputs
from .lookups_1d import get_segment_percentile_by_time
from .normalization import (
    NormalizationError,
    format_seconds,
    normalize_age_group,
    normalize_modality,
    normalize_sex_category,
    parse_segment_time_to_seconds,
)
from .sources import get_source


@lru_cache(maxsize=4)
def _read_cube_rows(path_text: str, sheet_name: str) -> tuple[dict[str, Any], ...]:
    workbook = load_workbook(Path(path_text), read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        rows: list[dict[str, Any]] = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            rows.append({str(column): value for column, value in zip(header, values)})
        return tuple(rows)
    finally:
        workbook.close()


def _cube_source_path() -> Path:
    source = get_source("sbr_cube")
    if isinstance(source.workbook, dict):
        raise ValueError("sbr_cube source must be a single workbook")
    return source.workbook


def _normalize_percentile(value: Any, name: str) -> float:
    percentile = float(value)
    if not 0 <= percentile <= 100:
        raise NormalizationError(f"{name} must be between 0 and 100")
    return percentile


def _cube_rows(modality: str, sex_label: str, age_group: str) -> list[dict[str, Any]]:
    rows = _read_cube_rows(str(_cube_source_path()), "Cube_Long")
    filtered = [
        row
        for row in rows
        if row.get("modality") == modality
        and row.get("sex_label") == sex_label
        and row.get("age_group") == age_group
    ]
    if not filtered:
        raise NormalizationError(f"No SBR cube rows found for {modality} {sex_label} {age_group}")
    return filtered


def _axis_rows(modality: str, sex_label: str, age_group: str) -> list[dict[str, Any]]:
    rows = _read_cube_rows(str(_cube_source_path()), "Cube_Axes")
    return [
        row
        for row in rows
        if row.get("modality") == modality
        and row.get("sex_label") == sex_label
        and row.get("age_group") == age_group
    ]


def _nearest_axis_time(axis_rows: list[dict[str, Any]], axis: str, percentile: float) -> float:
    candidates = [row for row in axis_rows if row.get("axis") == axis]
    if not candidates:
        raise NormalizationError(f"No axis rows found for {axis}")
    row = min(candidates, key=lambda item: abs(float(item["performance_percentile"]) - percentile))
    return float(row["seconds"])


def _bracket(values: list[float], target: float) -> tuple[float, float, str | None]:
    ordered = sorted(set(values))
    if target < ordered[0]:
        return ordered[0], ordered[0], "below_range"
    if target > ordered[-1]:
        return ordered[-1], ordered[-1], "above_range"
    for lower, upper in zip(ordered, ordered[1:]):
        if lower <= target <= upper:
            return lower, upper, None
    return ordered[-1], ordered[-1], "above_range"


def _lerp(a: float, b: float, t: float) -> float:
    return a + (b - a) * t


def _value_map(rows: list[dict[str, Any]]) -> dict[tuple[float, float, float], float]:
    return {
        (
            float(row["swim_performance_percentile"]),
            float(row["bike_performance_percentile"]),
            float(row["run_performance_percentile"]),
        ): float(row["joint_sbr_percentile"])
        for row in rows
    }


def _trilinear_interpolate(
    rows: list[dict[str, Any]],
    swim_percentile: float,
    bike_percentile: float,
    run_percentile: float,
) -> tuple[float, bool, str | None]:
    swims = [float(row["swim_performance_percentile"]) for row in rows]
    bikes = [float(row["bike_performance_percentile"]) for row in rows]
    runs = [float(row["run_performance_percentile"]) for row in rows]

    s0, s1, s_status = _bracket(swims, swim_percentile)
    b0, b1, b_status = _bracket(bikes, bike_percentile)
    r0, r1, r_status = _bracket(runs, run_percentile)
    if s_status or b_status or r_status:
        return float("nan"), False, s_status or b_status or r_status
    values = _value_map(rows)

    ts = 0 if s1 == s0 else (swim_percentile - s0) / (s1 - s0)
    tb = 0 if b1 == b0 else (bike_percentile - b0) / (b1 - b0)
    tr = 0 if r1 == r0 else (run_percentile - r0) / (r1 - r0)

    def v(s: float, b: float, r: float) -> float:
        return values[(s, b, r)]

    c000 = v(s0, b0, r0)
    c100 = v(s1, b0, r0)
    c010 = v(s0, b1, r0)
    c110 = v(s1, b1, r0)
    c001 = v(s0, b0, r1)
    c101 = v(s1, b0, r1)
    c011 = v(s0, b1, r1)
    c111 = v(s1, b1, r1)

    c00 = _lerp(c000, c100, ts)
    c10 = _lerp(c010, c110, ts)
    c01 = _lerp(c001, c101, ts)
    c11 = _lerp(c011, c111, ts)
    c0 = _lerp(c00, c10, tb)
    c1 = _lerp(c01, c11, tb)
    value = _lerp(c0, c1, tr)
    interpolated = (
        swim_percentile not in (s0, s1)
        or bike_percentile not in (b0, b1)
        or run_percentile not in (r0, r1)
    )
    return value, interpolated, s_status or b_status or r_status


def get_sbr_percentile_by_segment_percentiles(
    modality: Any,
    sex_category: Any,
    age_group: Any,
    swim_percentile: Any,
    bike_percentile: Any,
    run_percentile: Any,
) -> dict[str, Any]:
    modality = normalize_modality(modality)
    sex_label = normalize_sex_category(sex_category, field="sex_label")
    age_group = normalize_age_group(age_group)
    swim_percentile = _normalize_percentile(swim_percentile, "swim_percentile")
    bike_percentile = _normalize_percentile(bike_percentile, "bike_percentile")
    run_percentile = _normalize_percentile(run_percentile, "run_percentile")

    coverage = validate_query_inputs("sbr_cube", modality, sex_label, age_group)
    if not coverage.valid:
        return coverage.to_dict()

    rows = _cube_rows(modality, sex_label, age_group)
    axis_rows = _axis_rows(modality, sex_label, age_group)
    joint_percentile, interpolated, range_status = _trilinear_interpolate(
        rows,
        swim_percentile,
        bike_percentile,
        run_percentile,
    )
    if range_status is not None:
        return {
            "valid": False,
            "entity": "sbr_cube",
            "modality": modality,
            "sex_label": sex_label,
            "age_group": age_group,
            "swim_performance_percentile": swim_percentile,
            "bike_performance_percentile": bike_percentile,
            "run_performance_percentile": run_percentile,
            "range_status": range_status,
            "reason": "outside_empirical_range",
            "message": "At least one requested marginal percentile is outside the SBR cube's empirical grid.",
            "source": coverage.source,
            "sheet": coverage.sheet,
            "cube_resolution": coverage.details.get("cube_resolution"),
        }
    swim_seconds = _nearest_axis_time(axis_rows, "swim", swim_percentile)
    bike_seconds = _nearest_axis_time(axis_rows, "bike", bike_percentile)
    run_seconds = _nearest_axis_time(axis_rows, "run", run_percentile)

    return {
        "valid": True,
        "entity": "sbr_cube",
        "modality": modality,
        "sex_label": sex_label,
        "age_group": age_group,
        "swim_performance_percentile": swim_percentile,
        "bike_performance_percentile": bike_percentile,
        "run_performance_percentile": run_percentile,
        "swim_seconds": swim_seconds,
        "swim_time": format_seconds(swim_seconds),
        "bike_seconds": bike_seconds,
        "bike_time": format_seconds(bike_seconds),
        "run_seconds": run_seconds,
        "run_time": format_seconds(run_seconds),
        "joint_sbr_percentile": joint_percentile,
        "interpolated": interpolated,
        "range_status": range_status,
        "source": coverage.source,
        "sheet": coverage.sheet,
        "cube_resolution": coverage.details.get("cube_resolution"),
    }


def get_sbr_times_by_percentiles(
    modality: Any,
    sex_category: Any,
    age_group: Any,
    swim_percentile: Any,
    bike_percentile: Any,
    run_percentile: Any,
) -> dict[str, Any]:
    return get_sbr_percentile_by_segment_percentiles(
        modality,
        sex_category,
        age_group,
        swim_percentile,
        bike_percentile,
        run_percentile,
    )


def get_sbr_percentile_by_segment_times(
    modality: Any,
    sex_category: Any,
    age_group: Any,
    swim_time: Any,
    bike_time: Any,
    run_time: Any,
) -> dict[str, Any]:
    modality = normalize_modality(modality)
    sex_label = normalize_sex_category(sex_category, field="sex_label")
    age_group = normalize_age_group(age_group)

    coverage = validate_query_inputs("sbr_cube", modality, sex_label, age_group)
    if not coverage.valid:
        return coverage.to_dict()

    swim_seconds = parse_segment_time_to_seconds(modality, "Swim", swim_time)
    bike_seconds = parse_segment_time_to_seconds(modality, "Bike", bike_time)
    run_seconds = parse_segment_time_to_seconds(modality, "Run", run_time)
    swim_marginal = get_segment_percentile_by_time(modality, sex_label, age_group, "Swim", swim_seconds)
    bike_marginal = get_segment_percentile_by_time(modality, sex_label, age_group, "Bike", bike_seconds)
    run_marginal = get_segment_percentile_by_time(modality, sex_label, age_group, "Run", run_seconds)

    for marginal in (swim_marginal, bike_marginal, run_marginal):
        if not marginal.get("valid", True):
            return marginal

    joint = get_sbr_percentile_by_segment_percentiles(
        modality,
        sex_label,
        age_group,
        swim_marginal["performance_percentile"],
        bike_marginal["performance_percentile"],
        run_marginal["performance_percentile"],
    )
    if not joint.get("valid", False):
        return joint

    joint.update(
        {
            "input_swim_seconds": swim_seconds,
            "input_swim_time": format_seconds(swim_seconds),
            "input_bike_seconds": bike_seconds,
            "input_bike_time": format_seconds(bike_seconds),
            "input_run_seconds": run_seconds,
            "input_run_time": format_seconds(run_seconds),
            "swim_performance_percentile": swim_marginal["performance_percentile"],
            "bike_performance_percentile": bike_marginal["performance_percentile"],
            "run_performance_percentile": run_marginal["performance_percentile"],
            "swim_marginal_range_status": swim_marginal["range_status"],
            "bike_marginal_range_status": bike_marginal["range_status"],
            "run_marginal_range_status": run_marginal["range_status"],
            "method": "segment times -> marginal percentiles -> SBR cube",
        }
    )
    return joint
