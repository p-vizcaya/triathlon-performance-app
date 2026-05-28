from __future__ import annotations

import math
from functools import lru_cache
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook

from .normalization import format_seconds, normalize_sex_category
from .query_index import get_segment_params_from_index, get_total_params_from_index
from .sources import get_source


FINAL_REFERENCE = "Iteration 2"
CONFIDENCE_LEVEL = 0.95
Z_95 = 1.959963984540054
MIN_EVENT_PARAMS = 2


@lru_cache(maxsize=16)
def _read_sheet(path_text: str, sheet_name: str) -> tuple[dict[str, Any], ...]:
    workbook = load_workbook(Path(path_text), read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        rows: list[dict[str, Any]] = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            rows.append({str(column): value for column, value in zip(header, values)})
        return tuple(rows)
    finally:
        workbook.close()


@lru_cache(maxsize=16)
def _indexed_rows(path_text: str, sheet_name: str, key_columns: tuple[str, ...]) -> dict[tuple[Any, ...], tuple[dict[str, Any], ...]]:
    index: dict[tuple[Any, ...], list[dict[str, Any]]] = {}
    for row in _read_sheet(path_text, sheet_name):
        key = tuple(row.get(column) for column in key_columns)
        index.setdefault(key, []).append(row)
    return {key: tuple(values) for key, values in index.items()}


@lru_cache(maxsize=128)
def _read_matching_rows(
    path_text: str,
    sheet_name: str,
    criteria: tuple[tuple[str, Any], ...],
) -> tuple[dict[str, Any], ...]:
    workbook = load_workbook(Path(path_text), read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        columns = {str(column): index for index, column in enumerate(header)}
        wanted = [(column, value, columns[column]) for column, value in criteria if column in columns]
        rows: list[dict[str, Any]] = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            if all(values[index] == value for _, value, index in wanted):
                rows.append({str(column): value for column, value in zip(header, values)})
        return tuple(rows)
    finally:
        workbook.close()


def _source_path(entity: str, modality: str) -> Path:
    source = get_source(entity)
    workbook = source.workbook
    if isinstance(workbook, dict):
        return workbook[modality]
    return workbook


def _segment_reference_path() -> Path:
    public_path = _source_path("segment_curve", "Standard")
    compact_path = public_path.parent / "WT_Segment_Query_Source_1989_2025.xlsx"
    return compact_path if compact_path.exists() else public_path


def _percentile_ci(performance_percentile: float, n: int) -> tuple[float, float, float]:
    if n <= 0 or not np.isfinite(performance_percentile):
        return np.nan, np.nan, np.nan
    q_slow = 1 - performance_percentile / 100
    se = 100 * math.sqrt(max(0.0, q_slow * (1 - q_slow)) / n)
    lower = max(0.0, performance_percentile - Z_95 * se)
    upper = min(100.0, performance_percentile + Z_95 * se)
    return lower, upper, se


def _interpolate(points: list[tuple[float, float]], x: float) -> float:
    ordered = sorted(points, key=lambda point: point[0])
    xs = np.asarray([point[0] for point in ordered], dtype=float)
    ys = np.asarray([point[1] for point in ordered], dtype=float)
    return float(np.interp(float(x), xs, ys))


def _curve_by_seconds(points: list[tuple[float, float]]) -> list[tuple[float, float]]:
    return sorted(points, key=lambda point: point[0])


def _curve_by_percentile(points: list[tuple[float, float]]) -> list[tuple[float, float]]:
    return sorted(((percentile, seconds) for seconds, percentile in points), key=lambda point: point[0])


def _cumulative_params(rows: list[dict[str, Any]]) -> list[dict[str, float]]:
    by_event: dict[Any, list[dict[str, Any]]] = {}
    for row in rows:
        event = row.get("event_id", row.get("year"))
        by_event.setdefault(event, []).append(row)

    params = []
    for event_rows in by_event.values():
        alpha_c = 0.0
        beta_c = 1.0
        for row in sorted(event_rows, key=lambda item: int(item.get("iteration") or 0)):
            alpha = float(row["alpha"])
            beta = float(row["beta"])
            alpha_c = alpha + beta * alpha_c
            beta_c *= beta
        params.append({"alpha": alpha_c, "beta": beta_c})
    return params


def _total_param_rows(modality: str, sex_label: str, age_group: str) -> list[dict[str, Any]]:
    indexed = get_total_params_from_index(modality, sex_label, age_group)
    if indexed is not None:
        return indexed
    sex = normalize_sex_category(sex_label, field="sex")
    path = _source_path("total_time_curve", modality)
    return list(
        _read_matching_rows(
            str(path),
            "Event_Transform_Params",
            (("modality", modality), ("sex", sex), ("age_group", age_group)),
        )
    )


def _segment_param_rows(modality: str, sex_label: str, age_group: str, segment: str) -> list[dict[str, Any]]:
    indexed = get_segment_params_from_index(modality, sex_label, age_group, segment)
    if indexed is not None:
        return indexed
    path = _segment_reference_path()
    criteria = (
        ("modality", modality),
        ("sex_label", sex_label),
        ("age_group", age_group),
        ("segment", segment),
    )
    return list(
        _read_matching_rows(
            str(path),
            "Event_Transform_Params",
            criteria,
        )
    )


def _difficulty_times_from_percentile(seconds: float, params: list[dict[str, float]]) -> np.ndarray:
    if not np.isfinite(seconds) or seconds <= 0 or len(params) < MIN_EVENT_PARAMS:
        return np.array([], dtype=float)
    log_seconds = math.log(seconds)
    values = [
        math.exp((log_seconds - param["alpha"]) / param["beta"])
        for param in params
        if np.isfinite(param["alpha"]) and np.isfinite(param["beta"]) and abs(param["beta"]) > 1e-12
    ]
    return np.asarray([value for value in values if np.isfinite(value) and value > 0], dtype=float)


def _difficulty_percentiles_from_time(
    seconds: float,
    params: list[dict[str, float]],
    seconds_to_percentile: list[tuple[float, float]],
) -> np.ndarray:
    if not np.isfinite(seconds) or seconds <= 0 or len(params) < MIN_EVENT_PARAMS:
        return np.array([], dtype=float)
    ref_times = [
        math.exp(param["alpha"] + param["beta"] * math.log(seconds))
        for param in params
        if np.isfinite(param["alpha"]) and np.isfinite(param["beta"])
    ]
    return np.asarray([_interpolate(seconds_to_percentile, value) for value in ref_times if np.isfinite(value)], dtype=float)


def _time_summary(values: np.ndarray) -> dict[str, Any]:
    if len(values) < MIN_EVENT_PARAMS:
        return {}
    q10, q25, q50, q75, q90 = np.quantile(values, [0.10, 0.25, 0.50, 0.75, 0.90])
    return {
        "difficulty_time_q10_seconds": float(q10),
        "difficulty_time_q10": format_seconds(q10),
        "difficulty_time_q25_seconds": float(q25),
        "difficulty_time_q25": format_seconds(q25),
        "difficulty_time_median_seconds": float(q50),
        "difficulty_time_median": format_seconds(q50),
        "difficulty_time_q75_seconds": float(q75),
        "difficulty_time_q75": format_seconds(q75),
        "difficulty_time_q90_seconds": float(q90),
        "difficulty_time_q90": format_seconds(q90),
        "difficulty_time_iqr_width_seconds": float(q75 - q25),
    }


def _percentile_summary(values: np.ndarray) -> dict[str, Any]:
    if len(values) < MIN_EVENT_PARAMS:
        return {}
    q10, q25, q50, q75, q90 = np.quantile(values, [0.10, 0.25, 0.50, 0.75, 0.90])
    return {
        "difficulty_percentile_q10": float(q10),
        "difficulty_percentile_q25": float(q25),
        "difficulty_percentile_median": float(q50),
        "difficulty_percentile_q75": float(q75),
        "difficulty_percentile_q90": float(q90),
        "difficulty_percentile_iqr_width": float(q75 - q25),
    }


def uncertainty_for_time(
    *,
    modality: str,
    sex_label: str,
    age_group: str,
    points: list[tuple[float, float]],
    seconds: float,
    percentile: float,
    segment: str | None = None,
) -> dict[str, Any]:
    n = len(points)
    p_low, p_high, p_se = _percentile_ci(percentile, n)
    by_percentile = _curve_by_percentile(points)
    stat_fast = _interpolate(by_percentile, p_high)
    stat_slow = _interpolate(by_percentile, p_low)
    if segment is None:
        rows = _total_param_rows(modality, sex_label, age_group)
    else:
        rows = _segment_param_rows(modality, sex_label, age_group, segment)
    params = _cumulative_params(rows)
    difficulty_percentiles = _difficulty_percentiles_from_time(seconds, params, _curve_by_seconds(points))
    return {
        "uncertainty_confidence_level": CONFIDENCE_LEVEL,
        "stat_n": n,
        "event_param_count": len(params),
        "performance_percentile_se": p_se,
        "performance_percentile_ci_low": p_low,
        "performance_percentile_ci_high": p_high,
        "stat_time_fast_seconds": stat_fast,
        "stat_time_fast": format_seconds(stat_fast),
        "stat_time_slow_seconds": stat_slow,
        "stat_time_slow": format_seconds(stat_slow),
        "stat_time_width_seconds": stat_slow - stat_fast,
        **_percentile_summary(difficulty_percentiles),
    }


def uncertainty_for_percentile(
    *,
    modality: str,
    sex_label: str,
    age_group: str,
    points: list[tuple[float, float]],
    seconds: float,
    percentile: float,
    segment: str | None = None,
) -> dict[str, Any]:
    n = len(points)
    p_low, p_high, p_se = _percentile_ci(percentile, n)
    by_percentile = _curve_by_percentile(points)
    stat_fast = _interpolate(by_percentile, p_high)
    stat_slow = _interpolate(by_percentile, p_low)
    if segment is None:
        rows = _total_param_rows(modality, sex_label, age_group)
    else:
        rows = _segment_param_rows(modality, sex_label, age_group, segment)
    params = _cumulative_params(rows)
    difficulty_times = _difficulty_times_from_percentile(seconds, params)
    difficulty = _time_summary(difficulty_times)
    rec_fast = min(stat_fast, difficulty.get("difficulty_time_q25_seconds", stat_fast))
    rec_slow = max(stat_slow, difficulty.get("difficulty_time_q75_seconds", stat_slow))
    return {
        "uncertainty_confidence_level": CONFIDENCE_LEVEL,
        "stat_n": n,
        "event_param_count": len(params),
        "performance_percentile_se": p_se,
        "performance_percentile_ci_low": p_low,
        "performance_percentile_ci_high": p_high,
        "stat_time_fast_seconds": stat_fast,
        "stat_time_fast": format_seconds(stat_fast),
        "stat_time_slow_seconds": stat_slow,
        "stat_time_slow": format_seconds(stat_slow),
        "stat_time_width_seconds": stat_slow - stat_fast,
        **difficulty,
        "recommended_time_fast_seconds": rec_fast,
        "recommended_time_fast": format_seconds(rec_fast),
        "recommended_time_slow_seconds": rec_slow,
        "recommended_time_slow": format_seconds(rec_slow),
        "recommended_time_width_seconds": rec_slow - rec_fast,
    }
