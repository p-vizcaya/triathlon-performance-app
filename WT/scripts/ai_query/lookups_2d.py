from __future__ import annotations

from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .coverage import validate_query_inputs
from .lookups_1d import get_segment_percentile_by_time
from .normalization import (
    NormalizationError,
    format_seconds,
    normalize_age_group,
    normalize_modality,
    normalize_pair,
    normalize_sex_category,
    parse_segment_time_to_seconds,
    parse_time_to_seconds,
)
from .sources import get_source


@lru_cache(maxsize=4)
def _read_pair_rows(path_text: str) -> tuple[dict[str, Any], ...]:
    workbook = load_workbook(Path(path_text), read_only=True, data_only=True)
    try:
        worksheet = workbook["Pair_Performance_Tables"]
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        rows: list[dict[str, Any]] = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            rows.append({str(column): value for column, value in zip(header, values)})
        return tuple(rows)
    finally:
        workbook.close()


def _pair_source_path() -> Path:
    source = get_source("segment_pair_plane")
    if isinstance(source.workbook, dict):
        raise ValueError("segment_pair_plane source must be a single workbook")
    return source.workbook


def _normalize_percentile(value: Any, name: str = "percentile") -> float:
    percentile = float(value)
    if not 0 <= percentile <= 100:
        raise NormalizationError(f"{name} must be between 0 and 100")
    return percentile


def _pair_rows(modality: str, sex_label: str, age_group: str, pair: str) -> list[dict[str, Any]]:
    rows = _read_pair_rows(str(_pair_source_path()))
    filtered = [
        row
        for row in rows
        if row.get("modality") == modality
        and row.get("sex_label") == sex_label
        and row.get("age_group") == age_group
        and row.get("pair") == pair
    ]
    if not filtered:
        raise NormalizationError(f"No pair-plane rows found for {modality} {sex_label} {age_group} {pair}")
    return filtered


def _axis_segments(rows: list[dict[str, Any]]) -> tuple[str, str]:
    first = rows[0]
    return str(first["x_segment"]), str(first["y_segment"])


def _nearest_axis_times(rows: list[dict[str, Any]], x_percentile: float, y_percentile: float) -> tuple[float, float]:
    x_row = min(rows, key=lambda row: abs(float(row["x_performance_percentile"]) - x_percentile))
    y_row = min(rows, key=lambda row: abs(float(row["y_performance_percentile"]) - y_percentile))
    return float(x_row["x_seconds"]), float(y_row["y_seconds"])


def _cell_map(rows: list[dict[str, Any]]) -> dict[tuple[float, float], float]:
    return {
        (float(row["x_performance_percentile"]), float(row["y_performance_percentile"])): float(row["joint_pair_percentile"])
        for row in rows
    }


def _bracket(values: list[float], target: float) -> tuple[float, float, str | None]:
    ordered = sorted(set(values))
    if target <= ordered[0]:
        return ordered[0], ordered[0], "below_range"
    if target >= ordered[-1]:
        return ordered[-1], ordered[-1], "above_range"
    for lower, upper in zip(ordered, ordered[1:]):
        if lower <= target <= upper:
            return lower, upper, None
    return ordered[-1], ordered[-1], "above_range"


def _lerp(a: float, b: float, t: float) -> float:
    return a + (b - a) * t


def _bilinear_interpolate(rows: list[dict[str, Any]], x: float, y: float) -> tuple[float, bool, str | None]:
    xs = [float(row["x_performance_percentile"]) for row in rows]
    ys = [float(row["y_performance_percentile"]) for row in rows]
    x0, x1, x_status = _bracket(xs, x)
    y0, y1, y_status = _bracket(ys, y)
    values = _cell_map(rows)

    q00 = values[(x0, y0)]
    q01 = values[(x0, y1)]
    q10 = values[(x1, y0)]
    q11 = values[(x1, y1)]

    if x0 == x1 and y0 == y1:
        return q00, False, x_status or y_status
    if x0 == x1:
        t = 0 if y1 == y0 else (y - y0) / (y1 - y0)
        return _lerp(q00, q01, t), y not in (y0, y1), x_status or y_status
    if y0 == y1:
        t = 0 if x1 == x0 else (x - x0) / (x1 - x0)
        return _lerp(q00, q10, t), x not in (x0, x1), x_status or y_status

    tx = (x - x0) / (x1 - x0)
    ty = (y - y0) / (y1 - y0)
    lower = _lerp(q00, q10, tx)
    upper = _lerp(q01, q11, tx)
    return _lerp(lower, upper, ty), x not in (x0, x1) or y not in (y0, y1), x_status or y_status


def get_pair_percentile_by_segment_percentiles(
    modality: Any,
    sex_category: Any,
    age_group: Any,
    pair: Any,
    x_percentile: Any,
    y_percentile: Any,
) -> dict[str, Any]:
    modality = normalize_modality(modality)
    sex_label = normalize_sex_category(sex_category, field="sex_label")
    age_group = normalize_age_group(age_group)
    pair = normalize_pair(pair)
    x_percentile = _normalize_percentile(x_percentile, "x_percentile")
    y_percentile = _normalize_percentile(y_percentile, "y_percentile")

    coverage = validate_query_inputs("segment_pair_plane", modality, sex_label, age_group, pair=pair)
    if not coverage.valid:
        return coverage.to_dict()

    rows = _pair_rows(modality, sex_label, age_group, pair)
    x_segment, y_segment = _axis_segments(rows)
    joint_percentile, interpolated, range_status = _bilinear_interpolate(rows, x_percentile, y_percentile)
    x_seconds, y_seconds = _nearest_axis_times(rows, x_percentile, y_percentile)

    return {
        "valid": True,
        "entity": "segment_pair_plane",
        "modality": modality,
        "sex_label": sex_label,
        "age_group": age_group,
        "pair": pair,
        "x_segment": x_segment,
        "y_segment": y_segment,
        "x_performance_percentile": x_percentile,
        "y_performance_percentile": y_percentile,
        "x_seconds": x_seconds,
        "x_time": format_seconds(x_seconds),
        "y_seconds": y_seconds,
        "y_time": format_seconds(y_seconds),
        "joint_pair_percentile": joint_percentile,
        "interpolated": interpolated,
        "range_status": range_status,
        "source": coverage.source,
        "sheet": coverage.sheet,
    }


def get_pair_times_by_percentiles(
    modality: Any,
    sex_category: Any,
    age_group: Any,
    pair: Any,
    x_percentile: Any,
    y_percentile: Any,
) -> dict[str, Any]:
    return get_pair_percentile_by_segment_percentiles(
        modality,
        sex_category,
        age_group,
        pair,
        x_percentile,
        y_percentile,
    )


def get_pair_percentile_by_segment_times(
    modality: Any,
    sex_category: Any,
    age_group: Any,
    pair: Any,
    x_time: Any,
    y_time: Any,
) -> dict[str, Any]:
    modality = normalize_modality(modality)
    sex_label = normalize_sex_category(sex_category, field="sex_label")
    age_group = normalize_age_group(age_group)
    pair = normalize_pair(pair)

    coverage = validate_query_inputs("segment_pair_plane", modality, sex_label, age_group, pair=pair)
    if not coverage.valid:
        return coverage.to_dict()

    rows = _pair_rows(modality, sex_label, age_group, pair)
    x_segment, y_segment = _axis_segments(rows)
    x_seconds = parse_segment_time_to_seconds(modality, x_segment, x_time)
    y_seconds = parse_segment_time_to_seconds(modality, y_segment, y_time)
    x_marginal = get_segment_percentile_by_time(modality, sex_label, age_group, x_segment, x_seconds)
    y_marginal = get_segment_percentile_by_time(modality, sex_label, age_group, y_segment, y_seconds)

    if not x_marginal.get("entity") == "segment_curve":
        return x_marginal
    if not y_marginal.get("entity") == "segment_curve":
        return y_marginal

    joint = get_pair_percentile_by_segment_percentiles(
        modality,
        sex_label,
        age_group,
        pair,
        x_marginal["performance_percentile"],
        y_marginal["performance_percentile"],
    )
    if not joint.get("valid", False):
        return joint

    joint.update(
        {
            "input_x_seconds": x_seconds,
            "input_x_time": format_seconds(x_seconds),
            "input_y_seconds": y_seconds,
            "input_y_time": format_seconds(y_seconds),
            "x_performance_percentile": x_marginal["performance_percentile"],
            "y_performance_percentile": y_marginal["performance_percentile"],
            "x_marginal_range_status": x_marginal["range_status"],
            "y_marginal_range_status": y_marginal["range_status"],
            "method": "segment times -> marginal percentiles -> pair plane",
        }
    )
    return joint
