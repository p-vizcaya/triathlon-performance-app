from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .normalization import (
    NormalizationError,
    normalize_age_group,
    normalize_modality,
    normalize_pair,
    normalize_segment,
    normalize_sex_category,
)
from .sources import (
    PAIRS,
    SEGMENTS,
    STANDARD_AGE_GROUPS,
    SPRINT_AGE_GROUPS,
    SourceSpec,
    get_source,
)


@dataclass(frozen=True)
class CoverageResult:
    valid: bool
    entity: str
    modality: str
    sex_label: str
    age_group: str
    source: str | dict[str, str] | None = None
    sheet: str | None = None
    coverage_status: str | None = None
    records: int | None = None
    events: int | None = None
    reason: str | None = None
    message: str | None = None
    details: dict[str, Any] | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "valid": self.valid,
            "entity": self.entity,
            "modality": self.modality,
            "sex_label": self.sex_label,
            "age_group": self.age_group,
            "source": self.source,
            "sheet": self.sheet,
            "coverage_status": self.coverage_status,
            "records": self.records,
            "events": self.events,
            "reason": self.reason,
            "message": self.message,
            "details": self.details or {},
        }


def _public_age_groups(modality: str) -> tuple[str, ...]:
    return SPRINT_AGE_GROUPS if modality == "Sprint" else STANDARD_AGE_GROUPS


def _stringify_source(source: SourceSpec, modality: str) -> str | dict[str, str]:
    workbook = source.workbook
    if isinstance(workbook, dict):
        return {key: str(path) for key, path in workbook.items()} if modality not in workbook else str(workbook[modality])
    return str(workbook)


def _source_path(source: SourceSpec, modality: str) -> Path:
    workbook = source.workbook
    if isinstance(workbook, dict):
        return workbook[modality]
    return workbook


@lru_cache(maxsize=16)
def _read_sheet_rows(path_text: str, sheet_name: str) -> tuple[dict[str, Any], ...]:
    workbook = load_workbook(Path(path_text), read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        rows: list[dict[str, Any]] = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            rows.append({str(column): value for column, value in zip(header, values)})
        return tuple(rows)
    finally:
        workbook.close()


def _coverage_rows(source: SourceSpec, modality: str, sheet_name: str) -> tuple[dict[str, Any], ...]:
    return _read_sheet_rows(str(_source_path(source, modality)), sheet_name)


def _base_invalid(
    *,
    entity: str,
    modality: str,
    sex_label: str,
    age_group: str,
    reason: str,
    message: str,
    details: dict[str, Any] | None = None,
) -> CoverageResult:
    return CoverageResult(
        valid=False,
        entity=entity,
        modality=modality,
        sex_label=sex_label,
        age_group=age_group,
        reason=reason,
        message=message,
        details=details,
    )


def _validate_public_dimensions(entity: str, modality: str, sex_label: str, age_group: str) -> CoverageResult | None:
    if age_group not in _public_age_groups(modality):
        return _base_invalid(
            entity=entity,
            modality=modality,
            sex_label=sex_label,
            age_group=age_group,
            reason="age_group_not_exposed",
            message=f"{age_group} is not an exposed query age group for {modality}.",
            details={},
        )
    return None


def _built_result(
    *,
    source: SourceSpec,
    row: dict[str, Any],
    entity: str,
    modality: str,
    sex_label: str,
    age_group: str,
    sheet: str,
    details: dict[str, Any] | None = None,
) -> CoverageResult:
    return CoverageResult(
        valid=True,
        entity=entity,
        modality=modality,
        sex_label=sex_label,
        age_group=age_group,
        source=_stringify_source(source, modality),
        sheet=source.main_sheet,
        coverage_status=str(row.get("status", "Built")),
        records=int(row["records"]) if row.get("records") is not None else None,
        events=int(row["events"]) if row.get("events") is not None else None,
        details={"coverage_sheet": sheet, **(details or {})},
    )


def _not_found_result(
    *,
    source: SourceSpec,
    entity: str,
    modality: str,
    sex_label: str,
    age_group: str,
    reason: str = "coverage_not_found",
    message: str | None = None,
    details: dict[str, Any] | None = None,
) -> CoverageResult:
    return CoverageResult(
        valid=False,
        entity=entity,
        modality=modality,
        sex_label=sex_label,
        age_group=age_group,
        source=_stringify_source(source, modality),
        sheet=source.main_sheet,
        reason=reason,
        message=message or f"No {entity} coverage is available for {modality} {sex_label} {age_group}.",
        details=details,
    )


def _find_group_summary(source: SourceSpec, modality: str, sex_label: str, age_group: str) -> dict[str, Any] | None:
    sex = normalize_sex_category(sex_label, field="sex")
    for row in _coverage_rows(source, modality, "Group_Summary"):
        if row.get("modality") == modality and row.get("sex") == sex and row.get("age_group") == age_group:
            return row
    return None


def _validate_total_time_curve(source: SourceSpec, modality: str, sex_label: str, age_group: str) -> CoverageResult:
    row = _find_group_summary(source, modality, sex_label, age_group)
    if row is None:
        return _not_found_result(source=source, entity="total_time_curve", modality=modality, sex_label=sex_label, age_group=age_group)
    if row.get("status") != "Built":
        return _not_found_result(
            source=source,
            entity="total_time_curve",
            modality=modality,
            sex_label=sex_label,
            age_group=age_group,
            reason="insufficient_data",
            message=f"Total-time coverage is not sufficient for {modality} {sex_label} {age_group}.",
            details={"coverage_status": row.get("status")},
        )
    return _built_result(
        source=source,
        row=row,
        entity="total_time_curve",
        modality=modality,
        sex_label=sex_label,
        age_group=age_group,
        sheet="Group_Summary",
    )


def _validate_segment_curve(
    source: SourceSpec,
    modality: str,
    sex_label: str,
    age_group: str,
    segment: str | None,
) -> CoverageResult:
    if segment is None:
        return _base_invalid(
            entity="segment_curve",
            modality=modality,
            sex_label=sex_label,
            age_group=age_group,
            reason="missing_segment",
            message="Segment is required for segment-curve validation.",
        )
    segment = normalize_segment(segment)
    if segment not in SEGMENTS:
        return _base_invalid(
            entity="segment_curve",
            modality=modality,
            sex_label=sex_label,
            age_group=age_group,
            reason="invalid_segment",
            message=f"{segment} is not a supported segment.",
            details={"available_segments": list(SEGMENTS)},
        )
    row = _find_group_summary(source, modality, sex_label, age_group)
    if row is None or row.get("status") != "Built":
        return _not_found_result(
            source=source,
            entity="segment_curve",
            modality=modality,
            sex_label=sex_label,
            age_group=age_group,
            reason="insufficient_data",
            message=f"Segment coverage is not available for {modality} {sex_label} {age_group}.",
            details={"segment": segment, "coverage_status": None if row is None else row.get("status")},
        )
    return _built_result(
        source=source,
        row=row,
        entity="segment_curve",
        modality=modality,
        sex_label=sex_label,
        age_group=age_group,
        sheet="Group_Summary",
        details={"segment": segment},
    )


def _validate_pair_plane(
    source: SourceSpec,
    modality: str,
    sex_label: str,
    age_group: str,
    pair: str | None,
) -> CoverageResult:
    if pair is None:
        return _base_invalid(
            entity="segment_pair_plane",
            modality=modality,
            sex_label=sex_label,
            age_group=age_group,
            reason="missing_pair",
            message="Pair is required for segment-pair validation.",
        )
    pair = normalize_pair(pair)
    if pair not in PAIRS:
        return _base_invalid(
            entity="segment_pair_plane",
            modality=modality,
            sex_label=sex_label,
            age_group=age_group,
            reason="invalid_pair",
            message=f"{pair} is not a supported pair.",
            details={"available_pairs": list(PAIRS)},
        )

    for row in _coverage_rows(source, modality, "Pair_Summary"):
        if (
            row.get("modality") == modality
            and row.get("sex_label") == sex_label
            and row.get("age_group") == age_group
            and row.get("pair") == pair
            and row.get("status") == "Built"
        ):
            return _built_result(
                source=source,
                row=row,
                entity="segment_pair_plane",
                modality=modality,
                sex_label=sex_label,
                age_group=age_group,
                sheet="Pair_Summary",
                details={"pair": pair, "x_segment": row.get("x_segment"), "y_segment": row.get("y_segment")},
            )

    return _not_found_result(
        source=source,
        entity="segment_pair_plane",
        modality=modality,
        sex_label=sex_label,
        age_group=age_group,
        reason="coverage_not_found",
        message=f"No {pair} plane is available for {modality} {sex_label} {age_group}.",
        details={"pair": pair},
    )


def _validate_sbr_cube(source: SourceSpec, modality: str, sex_label: str, age_group: str) -> CoverageResult:
    for row in _coverage_rows(source, modality, "Cube_Summary"):
        if (
            row.get("modality") == modality
            and row.get("sex_label") == sex_label
            and row.get("age_group") == age_group
            and row.get("status") == "Built"
        ):
            return _built_result(
                source=source,
                row=row,
                entity="sbr_cube",
                modality=modality,
                sex_label=sex_label,
                age_group=age_group,
                sheet="Cube_Summary",
                details={"cube_resolution": row.get("cube_resolution"), "cube_cells": row.get("cube_cells")},
            )

    for row in _coverage_rows(source, modality, "Skipped_Categories"):
        if row.get("modality") == modality and row.get("sex_label") == sex_label and row.get("age_group") == age_group:
            return _not_found_result(
                source=source,
                entity="sbr_cube",
                modality=modality,
                sex_label=sex_label,
                age_group=age_group,
                reason="insufficient_data",
                message=f"No SBR cube is available for {modality} {sex_label} {age_group}.",
                details={"coverage_status": row.get("status"), "records": row.get("records")},
            )

    return _not_found_result(source=source, entity="sbr_cube", modality=modality, sex_label=sex_label, age_group=age_group)


def validate_query_inputs(
    entity: str,
    modality: Any,
    sex_category: Any,
    age_group: Any,
    *,
    segment: Any | None = None,
    pair: Any | None = None,
) -> CoverageResult:
    modality = normalize_modality(modality)
    sex_label = normalize_sex_category(sex_category, field="sex_label")
    age_group = normalize_age_group(age_group)
    source = get_source(entity)

    public_error = _validate_public_dimensions(entity, modality, sex_label, age_group)
    if public_error is not None:
        return public_error

    if entity == "total_time_curve":
        return _validate_total_time_curve(source, modality, sex_label, age_group)
    if entity == "segment_curve":
        return _validate_segment_curve(source, modality, sex_label, age_group, segment)
    if entity == "segment_pair_plane":
        return _validate_pair_plane(source, modality, sex_label, age_group, pair)
    if entity == "sbr_cube":
        return _validate_sbr_cube(source, modality, sex_label, age_group)

    raise ValueError(f"Unsupported entity: {entity!r}")
