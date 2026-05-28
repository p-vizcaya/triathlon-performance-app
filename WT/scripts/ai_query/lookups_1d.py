from __future__ import annotations

from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .coverage import validate_query_inputs
from .normalization import (
    NormalizationError,
    format_seconds,
    normalize_age_group,
    normalize_modality,
    normalize_segment,
    normalize_sex_category,
    parse_segment_time_to_seconds,
    parse_time_to_seconds,
)
from .query_index import get_segment_curve_from_index, get_total_curve_from_index
from .sources import get_source
from .uncertainty import uncertainty_for_percentile, uncertainty_for_time


FINAL_REFERENCE = "Iteration 2"


@lru_cache(maxsize=16)
def _read_rows(path_text: str, sheet_name: str) -> tuple[dict[str, Any], ...]:
    workbook = load_workbook(Path(path_text), read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        rows: list[dict[str, Any]] = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            rows.append({str(column): value for column, value in zip(header, values)})
        return tuple(rows)
    finally:
        workbook.close()


@lru_cache(maxsize=128)
def _read_matching_rows(
    path_text: str,
    sheet_name: str,
    criteria: tuple[tuple[str, Any], ...],
) -> tuple[dict[str, Any], ...]:
    workbook = load_workbook(Path(path_text), read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        columns = {str(column): index for index, column in enumerate(header)}
        wanted = [(column, value, columns[column]) for column, value in criteria if column in columns]
        rows: list[dict[str, Any]] = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            if all(values[index] == value for _, value, index in wanted):
                rows.append({str(column): value for column, value in zip(header, values)})
        return tuple(rows)
    finally:
        workbook.close()


def _source_path(entity: str, modality: str) -> Path:
    source = get_source(entity)
    workbook = source.workbook
    if isinstance(workbook, dict):
        return workbook[modality]
    return workbook


def _segment_reference_path() -> Path:
    public_path = _source_path("segment_curve", "Standard")
    compact_path = public_path.parent / "WT_Segment_Query_Source_1989_2025.xlsx"
    return compact_path if compact_path.exists() else public_path


def _total_curve_rows(modality: str, sex_label: str, age_group: str) -> list[tuple[float, float]]:
    indexed = get_total_curve_from_index(modality, sex_label, age_group)
    if indexed is not None:
        return sorted(indexed, key=lambda point: point[0])
    sex = normalize_sex_category(sex_label, field="sex")
    path = _source_path("total_time_curve", modality)
    rows = _read_matching_rows(
        str(path),
        "Reference_Curves",
        (
            ("modality", modality),
            ("sex", sex),
            ("age_group", age_group),
            ("reference", FINAL_REFERENCE),
        ),
    )
    points = [
        (float(row["total_seconds"]), float(row["performance_percentile"]))
        for row in rows
    ]
    return sorted(points, key=lambda point: point[0])


def _segment_curve_rows(modality: str, sex_label: str, age_group: str, segment: str) -> list[tuple[float, float]]:
    indexed = get_segment_curve_from_index(modality, sex_label, age_group, segment)
    if indexed is not None:
        return sorted(indexed, key=lambda point: point[0])
    path = _segment_reference_path()
    sheet = "Segment_Reference_Curves"
    criteria = (
        ("modality", modality),
        ("sex_label", sex_label),
        ("age_group", age_group),
        ("segment", segment),
        ("reference", FINAL_REFERENCE),
    )
    rows = _read_matching_rows(
        str(path),
        sheet,
        criteria,
    )
    points = [
        (float(row["seconds"]), float(row["performance_percentile"]))
        for row in rows
    ]
    return sorted(points, key=lambda point: point[0])


def _interpolate_y(points: list[tuple[float, float]], x: float) -> tuple[float, bool, str | None]:
    if not points:
        raise NormalizationError("No curve points are available for interpolation")
    if len(points) == 1:
        return points[0][1], False, "single_point"

    if x <= points[0][0]:
        return points[0][1], False, "below_range"
    if x >= points[-1][0]:
        return points[-1][1], False, "above_range"

    for left, right in zip(points, points[1:]):
        x0, y0 = left
        x1, y1 = right
        if x0 <= x <= x1:
            if x1 == x0:
                return y0, False, None
            ratio = (x - x0) / (x1 - x0)
            return y0 + ratio * (y1 - y0), x not in (x0, x1), None

    return points[-1][1], False, "above_range"


def _interpolate_x(points: list[tuple[float, float]], percentile: float) -> tuple[float, bool, str | None]:
    percentile_points = sorted(((percentile_value, seconds) for seconds, percentile_value in points), key=lambda point: point[0])
    return _interpolate_y(percentile_points, percentile)


def _base_result(
    *,
    entity: str,
    modality: str,
    sex_label: str,
    age_group: str,
    source: str | dict[str, str] | None,
    sheet: str | None,
    interpolated: bool,
    range_status: str | None,
) -> dict[str, Any]:
    return {
        "entity": entity,
        "modality": modality,
        "sex_label": sex_label,
        "age_group": age_group,
        "source": source,
        "sheet": sheet,
        "interpolated": interpolated,
        "range_status": range_status,
    }


def _normalize_percentile(percentile: Any) -> float:
    value = float(percentile)
    if not 0 <= value <= 100:
        raise NormalizationError("Percentile must be between 0 and 100")
    return value


def get_total_percentile_by_time(
    modality: Any,
    sex_category: Any,
    age_group: Any,
    total_time: Any,
) -> dict[str, Any]:
    modality = normalize_modality(modality)
    sex_label = normalize_sex_category(sex_category, field="sex_label")
    age_group = normalize_age_group(age_group)
    total_seconds = parse_time_to_seconds(total_time)
    coverage = validate_query_inputs("total_time_curve", modality, sex_label, age_group)
    if not coverage.valid:
        return coverage.to_dict()

    percentile, interpolated, range_status = _interpolate_y(
        points := _total_curve_rows(modality, sex_label, age_group),
        total_seconds,
    )
    return {
        **_base_result(
            entity="total_time_curve",
            modality=modality,
            sex_label=sex_label,
            age_group=age_group,
            source=coverage.source,
            sheet=coverage.sheet,
            interpolated=interpolated,
            range_status=range_status,
        ),
        "input_total_seconds": total_seconds,
        "input_total_time": format_seconds(total_seconds),
        "performance_percentile": percentile,
        "uncertainty": uncertainty_for_time(
            modality=modality,
            sex_label=sex_label,
            age_group=age_group,
            points=points,
            seconds=total_seconds,
            percentile=percentile,
        ),
    }


def get_total_time_by_percentile(
    modality: Any,
    sex_category: Any,
    age_group: Any,
    percentile: Any,
) -> dict[str, Any]:
    modality = normalize_modality(modality)
    sex_label = normalize_sex_category(sex_category, field="sex_label")
    age_group = normalize_age_group(age_group)
    percentile_value = _normalize_percentile(percentile)
    coverage = validate_query_inputs("total_time_curve", modality, sex_label, age_group)
    if not coverage.valid:
        return coverage.to_dict()

    seconds, interpolated, range_status = _interpolate_x(
        points := _total_curve_rows(modality, sex_label, age_group),
        percentile_value,
    )
    return {
        **_base_result(
            entity="total_time_curve",
            modality=modality,
            sex_label=sex_label,
            age_group=age_group,
            source=coverage.source,
            sheet=coverage.sheet,
            interpolated=interpolated,
            range_status=range_status,
        ),
        "percentile": percentile_value,
        "total_seconds": seconds,
        "total_time": format_seconds(seconds),
        "uncertainty": uncertainty_for_percentile(
            modality=modality,
            sex_label=sex_label,
            age_group=age_group,
            points=points,
            seconds=seconds,
            percentile=percentile_value,
        ),
    }


def get_segment_percentile_by_time(
    modality: Any,
    sex_category: Any,
    age_group: Any,
    segment: Any,
    segment_time: Any,
) -> dict[str, Any]:
    modality = normalize_modality(modality)
    sex_label = normalize_sex_category(sex_category, field="sex_label")
    age_group = normalize_age_group(age_group)
    segment = normalize_segment(segment)
    seconds = parse_segment_time_to_seconds(modality, segment, segment_time)
    coverage = validate_query_inputs("segment_curve", modality, sex_label, age_group, segment=segment)
    if not coverage.valid:
        return coverage.to_dict()

    percentile, interpolated, range_status = _interpolate_y(
        points := _segment_curve_rows(modality, sex_label, age_group, segment),
        seconds,
    )
    return {
        **_base_result(
            entity="segment_curve",
            modality=modality,
            sex_label=sex_label,
            age_group=age_group,
            source=coverage.source,
            sheet=coverage.sheet,
            interpolated=interpolated,
            range_status=range_status,
        ),
        "segment": segment,
        "input_seconds": seconds,
        "input_time": format_seconds(seconds),
        "performance_percentile": percentile,
        "uncertainty": uncertainty_for_time(
            modality=modality,
            sex_label=sex_label,
            age_group=age_group,
            points=points,
            seconds=seconds,
            percentile=percentile,
            segment=segment,
        ),
    }


def get_segment_time_by_percentile(
    modality: Any,
    sex_category: Any,
    age_group: Any,
    segment: Any,
    percentile: Any,
) -> dict[str, Any]:
    modality = normalize_modality(modality)
    sex_label = normalize_sex_category(sex_category, field="sex_label")
    age_group = normalize_age_group(age_group)
    segment = normalize_segment(segment)
    percentile_value = _normalize_percentile(percentile)
    coverage = validate_query_inputs("segment_curve", modality, sex_label, age_group, segment=segment)
    if not coverage.valid:
        return coverage.to_dict()

    seconds, interpolated, range_status = _interpolate_x(
        points := _segment_curve_rows(modality, sex_label, age_group, segment),
        percentile_value,
    )
    return {
        **_base_result(
            entity="segment_curve",
            modality=modality,
            sex_label=sex_label,
            age_group=age_group,
            source=coverage.source,
            sheet=coverage.sheet,
            interpolated=interpolated,
            range_status=range_status,
        ),
        "segment": segment,
        "percentile": percentile_value,
        "seconds": seconds,
        "time": format_seconds(seconds),
        "uncertainty": uncertainty_for_percentile(
            modality=modality,
            sex_label=sex_label,
            age_group=age_group,
            points=points,
            seconds=seconds,
            percentile=percentile_value,
            segment=segment,
        ),
    }
